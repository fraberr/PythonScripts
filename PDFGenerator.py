from tkinter import Tk, Label, Entry, Button, StringVar
from tkinter import messagebox
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import utils
from reportlab.lib.units import inch
from tkinter.ttk import Combobox
from tkinter.font import BOLD, Font
import openpyxl
import shutil
import os
import sys

                                                    #--------- CREAZIONE LOGO E SCALING ---------#
                                   

def get_logo_path(logo_filename):
    script_dir = os.path.dirname(sys.argv[0])
    logo_path = os.path.join(script_dir, logo_filename)
    return logo_path

def scale_logo(logo_path, max_width, max_height):
    img = utils.ImageReader(logo_path)
    img_width, img_height = img.getSize()

    # Calculate the scaling factor for the logo
    width_ratio = max_width / img_width
    height_ratio = max_height / img_height
    scaling_factor = min(width_ratio, height_ratio)

    # Scale the logo dimensions
    scaled_width = img_width * scaling_factor
    scaled_height = img_height * scaling_factor

    return scaled_width, scaled_height


                                                    #--------- CREAZIONE FILE PDF ---------#


def generate_pdf(fornitore, cantiere, data, numero_odt, descrizione, costi, tipo_acquisto, title):
    # Create a folder based on the "fornitore" value if it doesn't exist
    fornitore_folder = os.path.join(os.getcwd(), "PDF", "Fornitore", fornitore.replace(' ', '_'))
    os.makedirs(fornitore_folder, exist_ok=True)

    # Create a folder based on the "cantiere" value if it doesn't exist
    cantiere_folder = os.path.join(os.getcwd(), "PDF", "Cantiere", cantiere.replace(' ', '_'))
    os.makedirs(cantiere_folder, exist_ok=True)

    # Set the file paths for the PDFs
    fornitore_pdf_path = os.path.join(fornitore_folder, f'{title}.pdf')
    cantiere_pdf_path = os.path.join(cantiere_folder, f'{title}.pdf')

    # Check if the PDF file already exists in the "fornitore" folder
    if os.path.exists(fornitore_pdf_path):
        messagebox.showerror("File Error", f"A file with the name '{title}.pdf' already exists in the 'fornitore' folder.")
        return None

    # Check if the PDF file already exists in the "cantiere" folder
    if os.path.exists(cantiere_pdf_path):
        messagebox.showerror("File Error", f"A file with the name '{title}.pdf' already exists in the 'cantiere' folder.")
        return None

    # Create the PDF using reportlab
    c = canvas.Canvas(fornitore_pdf_path, pagesize=letter)

    logo1_path = get_logo_path('image.png')
    logo2_path = get_logo_path('image2.jpeg')

    # Calculate logo positions and scale them down if necessary
    logo_margin = 0.5 * inch  # Margin from the top and sides
    logo_max_width = 1.5 * inch  # Maximum width for the logos
    logo_max_height = 0.75 * inch  # Maximum height for the logos

    # Draw the first logo on the top right
    logo1_width, logo1_height = scale_logo(logo1_path, logo_max_width, logo_max_height)
    logo1_x = letter[0] - logo1_width - logo_margin
    logo1_y = letter[1] - logo_margin - logo1_height
    c.drawInlineImage(logo1_path, logo1_x, logo1_y, width=logo1_width, height=logo1_height)

    # Draw the second logo on the top left
    logo2_width, logo2_height = scale_logo(logo2_path, logo_max_width, logo_max_height)
    logo2_x = logo_margin
    logo2_y = letter[1] - logo_margin - logo2_height
    c.drawInlineImage(logo2_path, logo2_x, logo2_y, width=logo2_width, height=logo2_height)

    # Add title
    title_font_size = 40  # Adjusts the font size 
    title_width = c.stringWidth(title, fontName='Times-Roman', fontSize=title_font_size)
    title_x = (letter[0] - title_width) / 2 + 1
    title_y = letter[1] - 2.5 * inch  # Adjusts the vertical position
    c.setFont("Times-Bold", title_font_size)
    c.drawCentredString(title_x, title_y, title)

    # Set font and font size for the rest of the text
    text_font_size = 20  # Adjusts the font size 
    c.setFont("Times-Roman", text_font_size)

    # Set initial position for the text fields
    text_x = 0.5 * inch  # Adjusts the horizontal position
    text_start_x = 1.5 * inch  # Adjusts the horizontal position
    text_y = title_y - 0.75 * inch  # Adjusts the vertical position 
    text_start_y = title_y - 1.50 * inch  # Adjusts the vertical position
    line_height = 0.80 * inch  # Adjusts the line height

    # Add the fields to the PDF
    
    fields = [
        ("Fornitore:", fornitore),
        ("Cantiere:", cantiere),
        ("Data:", data),
        ("Numero DDT:", numero_odt),
        ("Descrizione:", descrizione),
        ("Costi:", costi),
        ("Tipo di Acquisto:", tipo_acquisto),
    ]

    text_y = text_start_y
    for label, value in fields:
        text = f"{label}  \" {value} \""
        c.setFont("Times-Roman", text_font_size)
        c.drawString(text_start_x, text_y, text)
        text_y -= line_height

    # Add a green line at the bottom of the PDF
    green_line_y = 0.5 * inch
    green_line_color = "#006600"  
    red, green, blue = tuple(int(green_line_color[i:i + 2], 16) for i in (1, 3, 5))
    c.setStrokeColorRGB(red / 255, green / 255, blue / 255) 
    c.setLineWidth(5) 

    # Calculate the line positions to leave space from the horizontal borders
    line_start_x = 0.25 * inch
    line_end_x = letter[0] - 0.25 * inch

    c.line(line_start_x, green_line_y, line_end_x, green_line_y)

    # Save the PDF
    c.save()

    # Copy the PDF to the "cantiere" folder
    shutil.copyfile(fornitore_pdf_path, cantiere_pdf_path)

    # Show success messages
    messagebox.showinfo("Success", "PDF generated successfully.")
    messagebox.showinfo("Success", f"PDF saved in:\n\n{fornitore_pdf_path}\n\nand\n\n{cantiere_pdf_path}")


                                                #--------- CREAZIONE SPREADSHEET FORNITORE ---------#


    excel_file_path = os.path.join(fornitore_folder, f'{fornitore}_costi.xlsx')
    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Load workbook if it exists, otherwise create a new one
    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Creates the grid on the excel if it's a new file
    if not os.path.exists(excel_file_path):
        fields = ["Fornitore:", "Cantiere:", "Data:", "Numeri DDT:", "Descrizione:", "Tipo di Acquisti:", "Costi:", "TOTALE: "]
        for i, field in enumerate(fields):
            cell = sheet.cell(row=1, column=i+1)
            cell.value = field

    # Gets the last used row in the sheet
    last_row = sheet.max_row
    next_row = last_row + 1

    # Retrieve the previous totale value
    previous_totale = sheet.cell(row=last_row, column=8).value
    if previous_totale == "TOTALE: " or previous_totale is None:
        previous_totale = 0

    # Calculate the new totale value
    costi = float(costi)
    previous_totale = float(previous_totale) if previous_totale is not None else 0
    totale = costi + previous_totale

    # Adds new values to the next available row
    values = [fornitore, cantiere, data, numero_odt, descrizione, tipo_acquisto, costi, totale]
    for i, value in enumerate(values):
        cell = sheet.cell(row=next_row, column=i+1)
        cell.value = value

    # Delete the previous totale cell if it contains a value
    if sheet.cell(row=last_row, column=8).value and sheet.cell(row=last_row, column=8).value != "TOTALE: ":
        sheet.cell(row=last_row, column=8).value = None

    workbook.save(excel_file_path)


                                                #--------- CREAZIONE SPREADSHEET CANTIERE ---------#


    excel_file_path = os.path.join(cantiere_folder, f'{cantiere}_costi.xlsx')
    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Load workbook if it exists, otherwise create a new one
    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    # Creates the grid on the excel if it's a new file
    if not os.path.exists(excel_file_path):
        fields = ["Fornitore:", "Cantiere:", "Data:", "Numeri DDT:", "Descrizione:", "Tipo di Acquisti:", "Costi:", "TOTALE: "]
        for i, field in enumerate(fields):
            cell = sheet.cell(row=1, column=i+1)
            cell.value = field

    # Gets the last used row in the sheet
    last_row = sheet.max_row
    next_row = last_row + 1

    # Retrieve the previous totale value
    previous_totale = sheet.cell(row=last_row, column=8).value
    if previous_totale == "TOTALE: " or previous_totale is None:
        previous_totale = 0

    # Calculate the new totale value
    costi = float(costi)
    previous_totale = float(previous_totale) if previous_totale is not None else 0
    totale = costi + previous_totale

    # Adds new values to the next available row
    values = [fornitore, cantiere, data, numero_odt, descrizione, tipo_acquisto, costi, totale]
    for i, value in enumerate(values):
        cell = sheet.cell(row=next_row, column=i+1)
        cell.value = value

    # Delete the previous totale cell if it contains a value
    if sheet.cell(row=last_row, column=8).value and sheet.cell(row=last_row, column=8).value != "TOTALE: ":
        sheet.cell(row=last_row, column=8).value = None

    workbook.save(excel_file_path)
    
    return None


                                            #--------- CREAZIONE INTERFACCIA UTENTE GUI ---------#


def generate_pdf_gui():
    def generate_pdf_action():
        fornitore = fornitore_combobox.get()
        cantiere = cantiere_combobox.get()
        data = data_entry.get()
        numero_odt = numero_odt_var.get()
        descrizione = descrizione_var.get()    
        costi = costi_var.get()
        tipo_acquisto = tipo_acquisto_combobox.get()
        title = title_var.get()  # Get the title from the entry field

        if not title :
            messagebox.showerror("Title Error", "Inserisci un titolo")
        # Check for character limit
        elif len(fornitore) > 45 or len(cantiere) > 45 or len(numero_odt) > 45 or len(descrizione) > 45 or len(costi) > 45 or len(tipo_acquisto) > 45 or len(title) > 45:
            messagebox.showerror("Character Limit Exceeded", "Il limite massimo di caratteri è 50.")
        elif not costi :
            messagebox.showerror("Invalid Input", "Aggiungi un valore simbolico a \"Costi\"")
        else:
            pdf_paths = generate_pdf(fornitore, cantiere, data, numero_odt, descrizione, costi, tipo_acquisto, title)
            if pdf_paths is not None:
                fornitore_pdf_path, cantiere_pdf_path = pdf_paths
                output_label.config(text=f'PDFs generated successfully.\nFornitore PDF: {fornitore_pdf_path}\nCantiere PDF: {cantiere_pdf_path}')

    root = Tk()
    root.title("Generate PDF")
    root.geometry("500x600")

    # Increase the font size for labels and entry fields
    font_size = 18

    fornitore_label = Label(root, text="Fornitore:", font=("Times New Roman", font_size,))
    fornitore_label.pack()
    fornitore_combobox = Combobox(root, values=["EuropaElettronica", "Fogliani", "Carboni", "ElettroMeccanica2B", "CDM", "Comet", "Tecnomat", "Socfeder"], font=("Times New Roman", font_size))
    fornitore_combobox.config(height=200, width=20)  # Increase combobox height and width
    fornitore_combobox.pack()

    #No menu tendina
    cantiere_label = Label(root, text="Cantiere:", font=("Times New Roman", font_size))
    cantiere_label.pack()
    cantiere_combobox = Combobox(root, values=["Cantiere A", "Cantiere B", "Cantiere C", "Cantiere D", "Cantiere E", "Cantiere F"], font=("Times New Roman", font_size))
    cantiere_combobox.pack()

    data_label = Label(root, text="Data:", font=("Times New Roman", font_size))
    data_label.pack()
    data_var = StringVar()
    data_entry = Entry(root, textvariable=data_var, font=("Times New Roman", font_size))
    data_entry.pack()

    numero_odt_label = Label(root, text="Numero DDT:", font=("Times New Roman", font_size))
    numero_odt_label.pack()
    numero_odt_var = StringVar()
    numero_odt_entry = Entry(root, textvariable=numero_odt_var, font=("Times New Roman", font_size))
    numero_odt_entry.pack()

    descrizione_label = Label(root, text="Descrizione:", font=("Times New Roman", font_size))
    descrizione_label.pack()
    descrizione_var = StringVar()
    descrizione_entry = Entry(root, textvariable=descrizione_var, font=("Times New Roman", font_size))
    descrizione_entry.pack()

    costi_label = Label(root, text="Costi:", font=("Times New Roman", font_size))
    costi_label.pack()
    costi_var = StringVar()
    costi_entry = Entry(root, textvariable=costi_var, font=("Times New Roman", font_size))
    costi_entry.pack()

    tipo_acquisto_label = Label(root, text="Tipo Acquisto:", font=("Times New Roman", font_size))
    tipo_acquisto_label.pack()
    tipo_acquisto_combobox = Combobox(root, values=["MC", "P1", "P2"], font=("Times New Roman", font_size))
    tipo_acquisto_combobox.pack()

    #NO titolo, ma funzione ?
    title_label = Label(root, text="Titolo:", font=("Times New Roman", font_size))
    title_label.pack()
    title_var = StringVar()
    title_entry = Entry(root, textvariable=title_var, font=("Times New Roman", font_size))
    title_entry.pack()

    generate_button = Button(root, text="Generate PDF", command=generate_pdf_action, font=("Times New Roman", font_size))
    generate_button.pack()

    output_label = Label(root, text="", font=("Times New Roman", font_size))
    output_label.pack()

    root.mainloop()


generate_pdf_gui()